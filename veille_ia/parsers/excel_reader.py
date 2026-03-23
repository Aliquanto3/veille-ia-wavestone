"""Lecteur Excel pour les fichiers de veille IA (onglets NEWS + _DATA)."""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from veille_ia.config import CATEGORIES, SUB_CATEGORIES
from veille_ia.models import NewsItem

logger = logging.getLogger(__name__)


class ExcelReader:
    """Lit les news depuis un fichier Excel structuré (onglets NEWS + _DATA)."""

    EXPECTED_COLUMNS = ["Catégorie", "Sous-catégorie", "Titre", "Description", "Sources", "Date"]

    def __init__(self, file_path: Path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("Le module 'openpyxl' est requis pour lire les fichiers Excel.") from exc

        if not file_path.exists():
            raise FileNotFoundError(f"Fichier Excel introuvable : {file_path}")

        self.file_path = file_path
        self.wb = load_workbook(file_path, read_only=True, data_only=True)

    def __enter__(self) -> ExcelReader:
        return self

    def __exit__(self, *exc: Any) -> None:
        self.close()

    def read_vocabulary(self) -> Tuple[List[str], List[str]]:
        """Lit les catégories et sous-catégories depuis l'onglet _DATA."""
        if "_DATA" not in self.wb.sheetnames:
            logger.warning("Onglet _DATA absent — utilisation du vocabulaire par défaut.")
            return CATEGORIES[:], SUB_CATEGORIES[:]

        ws = self.wb["_DATA"]
        categories: List[str] = []
        sub_categories: List[str] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip():
                categories.append(str(row[0]).strip())
            if len(row) > 1 and row[1] and str(row[1]).strip():
                sub_categories.append(str(row[1]).strip())

        if not categories:
            logger.warning("Aucune catégorie dans _DATA — vocabulaire par défaut.")
            categories = CATEGORIES[:]
        if not sub_categories:
            logger.warning("Aucune sous-catégorie dans _DATA — vocabulaire par défaut.")
            sub_categories = SUB_CATEGORIES[:]

        return categories, sub_categories

    def read_news(self) -> List[NewsItem]:
        """Lit toutes les lignes de l'onglet NEWS et retourne des NewsItem."""
        if "NEWS" not in self.wb.sheetnames:
            raise ValueError(
                f"Onglet 'NEWS' introuvable dans {self.file_path}. "
                f"Onglets disponibles : {self.wb.sheetnames}"
            )

        ws = self.wb["NEWS"]
        col_map = self._build_column_map(ws)
        items: List[NewsItem] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            item = self._parse_row(row, col_map)
            if item:
                items.append(item)

        logger.info("📊 Excel : %d news lues depuis l'onglet NEWS.", len(items))
        return items

    def _build_column_map(self, ws: Any) -> Dict[str, int]:
        """Mappe les en-têtes de colonnes vers leurs indices."""
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else "" for h in header_row]
        col_map = {h: i for i, h in enumerate(headers)}

        missing = [c for c in ["Catégorie", "Titre"] if c not in col_map]
        if missing:
            raise ValueError(f"Colonnes obligatoires manquantes dans NEWS : {missing}")
        return col_map

    def _parse_row(self, row: tuple, col_map: Dict[str, int]) -> Optional[NewsItem]:
        """Parse une ligne Excel en NewsItem (ou None si ligne vide)."""
        title_val = self._get_cell(row, col_map, "Titre")
        if not title_val or not str(title_val).strip():
            return None

        return NewsItem(
            category=self._get_cell_str(row, col_map, "Catégorie", "Général"),
            sub_category=self._get_cell_str(row, col_map, "Sous-catégorie", "Général"),
            title=str(title_val).strip(),
            description=self._get_cell_str(row, col_map, "Description", ""),
            links=self._parse_sources(self._get_cell(row, col_map, "Sources")),
            date=self._parse_date(self._get_cell(row, col_map, "Date")),
        )

    @staticmethod
    def _get_cell(row: tuple, col_map: Dict[str, int], col_name: str) -> Any:
        idx = col_map.get(col_name)
        if idx is not None and idx < len(row):
            return row[idx]
        return None

    @staticmethod
    def _get_cell_str(row: tuple, col_map: Dict[str, int], col_name: str, default: str = "") -> str:
        idx = col_map.get(col_name)
        if idx is not None and idx < len(row) and row[idx]:
            return str(row[idx]).strip()
        return default

    @staticmethod
    def _parse_sources(raw: Any) -> List[str]:
        """Parse les URLs depuis une cellule (séparées par \\n ou ;)."""
        if not raw:
            return []
        links: List[str] = []
        for part in re.split(r"[\n;]+", str(raw).strip()):
            url = part.strip()
            if url.startswith("http"):
                links.append(url)
        return links

    @staticmethod
    def _parse_date(raw: Any) -> Optional[str]:
        """Convertit une valeur de cellule en string YYYY-MM-DD."""
        if not raw:
            return None
        if isinstance(raw, datetime):
            return raw.strftime("%Y-%m-%d")
        return str(raw).strip()[:10]

    def close(self) -> None:
        self.wb.close()
